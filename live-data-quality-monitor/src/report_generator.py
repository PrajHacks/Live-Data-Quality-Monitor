"""Generate a multi-sheet Excel data-quality report."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


Issue = dict[str, Any]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)
LABEL_FONT = Font(bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="B7B7B7"))

STATUS_FILLS = {
    "Excellent": PatternFill(fill_type="solid", fgColor="D9EAD3"),
    "Warning": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "Failed": PatternFill(fill_type="solid", fgColor="F4CCCC"),
}

SEVERITY_FILLS = {
    "Critical": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "Warning": PatternFill(fill_type="solid", fgColor="FCE5CD"),
    "Info": PatternFill(fill_type="solid", fgColor="CFE2F3"),
}

SEVERITY_ORDER = {
    "Critical": 0,
    "Warning": 1,
    "Info": 2,
}


def _normalized_series(series: pd.Series) -> pd.Series:
    """Normalize text-like values for column analysis."""

    if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
        return series

    normalized = series.astype("string").str.strip()
    return normalized.replace("", pd.NA)


def _set_cell_style(cell, *, bold: bool = False, fill: PatternFill | None = None, align: Alignment | None = None) -> None:
    """Apply the common cell styling used throughout the workbook."""

    cell.font = Font(bold=bold)
    if fill is not None:
        cell.fill = fill
    cell.border = THIN_BORDER
    if align is not None:
        cell.alignment = align


def _style_header_row(ws, row: int = 1) -> None:
    """Bold and fill a header row."""

    for cell in ws[row]:
        if cell.value is not None:
            _set_cell_style(
                cell,
                bold=True,
                fill=HEADER_FILL,
                align=Alignment(horizontal="center", vertical="center"),
            )


def _auto_adjust_column_widths(
    ws,
    *,
    min_width: int = 10,
    max_width: int = 60,
    ignore_rows: set[int] | None = None,
) -> None:
    """Auto-size columns based on displayed content with a practical cap."""

    ignore_rows = ignore_rows or set()

    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_len = len(str(column_cells[0].value)) if column_cells[0].value is not None else 0

        for cell in column_cells:
            if cell.row in ignore_rows or cell.value is None:
                continue

            value = cell.value
            if isinstance(value, datetime):
                text = value.strftime("%Y-%m-%d %H:%M")
            else:
                text = str(value)
            max_len = max(max_len, len(text))

        ws.column_dimensions[column_letter].width = min(max(max_len + 2, min_width), max_width)


def _sort_issues(issues: list[Issue]) -> list[Issue]:
    """Sort issues by severity and then by column for stable presentation."""

    return sorted(
        issues,
        key=lambda issue: (
            SEVERITY_ORDER.get(issue.get("severity"), 99),
            str(issue.get("column", "")),
            str(issue.get("issue_type", "")),
        ),
    )


def _write_title_banner(ws, title: str, status: str) -> None:
    """Write a filled title banner at the top of the summary sheet."""

    fill = STATUS_FILLS.get(status, STATUS_FILLS["Warning"])
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    for column in range(1, 5):
        ws.cell(row=1, column=column).fill = fill
        ws.cell(row=1, column=column).border = THIN_BORDER


def _write_summary_sheet(ws, df: pd.DataFrame, score_result: dict[str, Any]) -> None:
    """Populate the summary sheet with the score block and report metadata."""

    status = str(score_result.get("status", "Warning"))
    report_title = (
        "Data Quality Report"
        if str(score_result.get("analysis_mode")) == "generic"
        else "Open Food Facts Data Quality Report"
    )
    _write_title_banner(ws, report_title, status)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A2"] = "Generated at"
    ws["B2"] = generated_at
    ws["A3"] = "Rows analyzed"
    ws["B3"] = len(df)

    for label_cell in ("A2", "A3"):
        ws[label_cell].font = LABEL_FONT
        ws[label_cell].alignment = Alignment(horizontal="left")
    for value_cell in ("B2", "B3"):
        ws[value_cell].alignment = Alignment(horizontal="left")

    start_row = 5
    headers = ["Metric", "Value"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        _set_cell_style(
            cell,
            bold=True,
            fill=HEADER_FILL,
            align=Alignment(horizontal="center", vertical="center"),
        )

    score_rows = [
        ("Completeness", score_result.get("completeness", 0.0)),
        ("Validity", score_result.get("validity", 0.0)),
        ("Uniqueness", score_result.get("uniqueness", 0.0)),
        ("Consistency", score_result.get("consistency", 0.0)),
        ("Overall Score", score_result.get("overall_score", 0.0)),
        ("Status", status),
    ]

    for offset, (label, value) in enumerate(score_rows, start=1):
        row_index = start_row + offset
        label_cell = ws.cell(row=row_index, column=1, value=label)
        value_cell = ws.cell(row=row_index, column=2, value=value)
        _set_cell_style(label_cell, bold=True, align=Alignment(horizontal="left"))

        if label == "Status":
            fill = STATUS_FILLS.get(status, STATUS_FILLS["Warning"])
            _set_cell_style(
                value_cell,
                bold=True,
                fill=fill,
                align=Alignment(horizontal="left"),
            )
        else:
            _set_cell_style(value_cell, align=Alignment(horizontal="left"))
            value_cell.number_format = "0.0%"
            value_cell.value = float(value) / 100.0

    ws.freeze_panes = "A6"
    _auto_adjust_column_widths(ws, ignore_rows={1})


def _write_issues_sheet(ws, issues: list[Issue]) -> None:
    """Populate the issue log sheet."""

    headers = ["column", "issue_type", "count", "severity", "details"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _set_cell_style(
            cell,
            bold=True,
            fill=HEADER_FILL,
            align=Alignment(horizontal="center", vertical="center"),
        )

    for row_idx, issue in enumerate(_sort_issues(issues), start=2):
        row_values = [
            issue.get("column"),
            issue.get("issue_type"),
            issue.get("count"),
            issue.get("severity"),
            issue.get("details"),
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 4:
                fill = SEVERITY_FILLS.get(str(value), None)
                _set_cell_style(
                    cell,
                    bold=True,
                    fill=fill,
                    align=Alignment(horizontal="center", vertical="top"),
                )
            else:
                _set_cell_style(cell, align=Alignment(vertical="top", wrap_text=True))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_adjust_column_widths(ws, max_width=80)


def _write_column_analysis_sheet(ws, df: pd.DataFrame) -> None:
    """Populate the column profiling sheet."""

    headers = ["column_name", "data_type", "missing_count", "missing_pct", "unique_value_count"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _set_cell_style(
            cell,
            bold=True,
            fill=HEADER_FILL,
            align=Alignment(horizontal="center", vertical="center"),
        )

    total_rows = len(df)
    for row_idx, column in enumerate(df.columns, start=2):
        series = df[column]
        normalized = _normalized_series(series)
        missing_count = int(normalized.isna().sum())
        missing_pct = (missing_count / total_rows * 100.0) if total_rows else 0.0
        unique_count = int(normalized.dropna().nunique())

        values = [
            column,
            str(series.dtype),
            missing_count,
            missing_pct / 100.0,
            unique_count,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 4:
                cell.number_format = "0.0%"
            _set_cell_style(cell, align=Alignment(horizontal="left"))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_adjust_column_widths(ws)


def _write_sample_sheet(ws, df: pd.DataFrame) -> None:
    """Populate the sheet with the first 100 rows of the dataset."""

    sample_df = df.head(100)
    for col_idx, column in enumerate(sample_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        _set_cell_style(
            cell,
            bold=True,
            fill=HEADER_FILL,
            align=Alignment(horizontal="center", vertical="center"),
        )

    for row_idx, row in enumerate(sample_df.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _set_cell_style(cell, align=Alignment(vertical="top", wrap_text=True))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _auto_adjust_column_widths(ws, max_width=45)


def generate_excel_report(
    df: pd.DataFrame,
    issues: list[Issue],
    score_result: dict[str, Any],
    output_path: str | Path,
) -> Path:
    """Create the multi-sheet Excel report and save it to disk."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    issues_ws = workbook.create_sheet("Issues")
    column_analysis_ws = workbook.create_sheet("Column Analysis")
    sample_ws = workbook.create_sheet("Sample Data")

    _write_summary_sheet(summary_ws, df, score_result)
    _write_issues_sheet(issues_ws, issues)
    _write_column_analysis_sheet(column_analysis_ws, df)
    _write_sample_sheet(sample_ws, df)

    workbook.save(output_path)
    return output_path
